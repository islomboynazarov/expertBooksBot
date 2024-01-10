from telegram.error import TelegramError
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram import (ReplyKeyboardMarkup, ReplyKeyboardRemove)
import datetime
import sqlite3
import random
import telegram.ext
from datetime import date
from datetime import datetime, timedelta
from datetime import datetime
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, ConversationHandler,CallbackQueryHandler
from datetime import datetime as er
import xlwt
from openpyxl import load_workbook
import sqlite3 as sql
import os
from sqlite3 import Error
from telegram import KeyboardButton
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram import (ReplyKeyboardMarkup, ReplyKeyboardRemove)
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
import re
heelp=['admin wil update']
import logging

wallo=['bc1qnjh086d833wkdfs5z36x72fxuwjysgtyks0fqr']

man=[]

from telegram import LabeledPrice, ShippingOption
from telegram.ext import (
    Updater,
    CommandHandler,
    MessageHandler,
    Filters,
    PreCheckoutQueryHandler,
    ShippingQueryHandler,
)
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)
wc=['www']
import ast
import requests
from telegram import ParseMode
AB,PPR,CATE,DESCRIP,IMGE,BUTTON,DELETE,MCAT,DELCAT,JE,LOG,PDB,PDM,ADRS,PTYPE,WALL,ABD,DELA,FFAQ,HAM,PENORD,AMAR,PAY,KUNG,KAKAK=range(25)
def get_transaction_details(transaction_hash):
    api_url = f'https://blockchain.info/rawtx/{transaction_hash}'
    
    try:
        response = requests.get(api_url)
        response.raise_for_status()  # Check if the request was successful
        transaction_data = response.json()
        return transaction_data
    except requests.exceptions.RequestException as e:
        print(f"Error fetching transaction details: {e}")
        return None

def get_current_btc_to_usd_rate():
    # Use CoinGecko API to get the current BTC to USD exchange rate
    try:
        response = requests.get('https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=eur')
        response.raise_for_status()
        data = response.json()
        return data['bitcoin']['eur']
    except requests.exceptions.RequestException as e:
        print(f"Error fetching BTC to USD rate: {e}")
        return None
def usd_to_btc(amount_usd):
    api_url = 'https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=eur'
    
    try:
        response = requests.get(api_url)
        data = response.json()
        btc_to_usd_rate = data['bitcoin']['eur']
        amount_btc = amount_usd / btc_to_usd_rate
        return amount_btc
    
    except Exception as e:
        print(f"Error: {e}")
        return None
def start(update, context):
    print(update.effective_user.id)
    user = update.message.from_user
    usa=str(update.effective_user.id)
    userg = update.message.from_user
    xg=er.today().strftime("%m/%d/%Y, %H:%M:%S")
    try:
        usaf=userg.username
    except:
        usaf=userg.first_name#6417933558
  

    connection = sqlite3.connect("wallet.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT id FROM COMPANY where id= {}".format(int(update.effective_user.id))) 
    jobs = cursor.fetchall()
    if len(jobs) ==0: 
                cursor.execute("INSERT INTO COMPANY (ID,balance ,link,code,pid,ppr,ct,amount,payment,tref,refby,sub,name,allow,uname,address,number,post,province,street,note,tome) \
                    VALUES ({}, '{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".format(int(update.effective_user.id),"0","0","0","0","0","0","0","0","0","0","0",usaf,"0","0","0","0","0","0","0","0",xg))
                connection.commit()
                connection.close() 
                conn = sqlite3.connect('cart.db') 
                conn.execute("INSERT INTO COMPANY (ID,name,productID,price) \
                        VALUES ('{}', '{}','{}', '{}')".format(str(update.effective_user.id),'0','0','0'))
                conn.commit()
    loc_keyboard1 = KeyboardButton(text="🛒 Store")
    loc_keyboard2 = KeyboardButton(text="💵 Wallet")
    loc_keyboard3 = KeyboardButton(text="🗓 Orders")
    loc_keyboard4 = KeyboardButton(text="☎️ Support")
    keyboard = [[loc_keyboard1,loc_keyboard3],[loc_keyboard2,loc_keyboard4]]
    reply_markup = ReplyKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_photo(chat_id=update.effective_user.id,photo=open('logo.jpg','rb'),caption=wc[0],reply_markup=reply_markup)
    return BUTTON 
def button(update,context):
    h=update.effective_user.id
    query = update.callback_query
    a=query.data
    if a== '1':
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send name of product',reply_markup=reply_markup)
        return AB
    elif a=='200d':
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT ct from COMPANY where ID= '{}' ".format(update.effective_user.id))
        conn.commit()
        for row in cursor:
            msg=row[0]
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT name from COMPANY where category= '{}' ".format(msg))  
        conn.commit()                             
        keyf=[]
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('shop.db')
            cursor = conn.execute("SELECT name from COMPANY where category= '{}' ".format(msg))
            conn.commit()
            for row in cursor:
                c=[InlineKeyboardButton(row[0], callback_data=row[0])]
                keyf.append(c)
            b=[InlineKeyboardButton("Back", callback_data='🌎Main Menu')]
            keyf.append(b)
            reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text="🛒 Shop\n\nSelect Name of Product to View ",reply_markup=reply_markup)            
            return PDB
    elif a=='2f':
        conn = sqlite3.connect('faq.db')
        cursor = conn.execute("SELECT question,code from COMPANY ")
        conn.commit()
        keyf=[]
        for row in cursor:
            c=[InlineKeyboardButton(row[0], callback_data=row[1])]
            keyf.append(c) 
        b=[InlineKeyboardButton("🌎Main Menu", callback_data='🌎Main Menu')]
        keyf.append(b) 
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True ) 
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Queston to Delete",reply_markup=reply_markup)
        return DELA
    elif a=='2':
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send ProductID to delete',reply_markup=reply_markup)
        return DELETE
    elif a=='32':   
            keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send name of category',reply_markup=reply_markup)
            return MCAT
    elif a== '1f':
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Send question of FAQs",reply_markup=reply_markup)
        return ABD
    elif a== '99v':
        query.answer()  
        keyboard =[[InlineKeyboardButton("Pending", callback_data="Pending"),InlineKeyboardButton("History", callback_data="History")],[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='Select order type',reply_markup=reply_markup)
        return HAM
    elif a=='buy':
        cc=update.callback_query.message.caption
        fg=cc
        cc=cc.replace("$","")
        cf=update.callback_query.message.message_id
        dd=cc.split("𝐒𝐞𝐫𝐯𝐢𝐜𝐞 𝐍𝐚𝐦𝐞: ")
        dd=dd[1]
        dd=dd.split("𝐏𝐫𝐨𝐝𝐮𝐜𝐭𝐈𝐃: ")
        dd=dd[0]
        dd=dd.strip()

        ccv=cc.split("𝐏𝐫𝐨𝐝𝐮𝐜𝐭𝐈𝐃: ")
        ccv=ccv[1]
        ccv=ccv.split("𝐏𝐫𝐢𝐜𝐞: ")
        ccv=ccv[0]
        ccv=ccv.strip() 

        rr=cc.split("𝐏𝐫𝐢𝐜𝐞: ")
        rr=rr[1]
        rr=rr.split("𝐃𝐞𝐬𝐜𝐫𝐢𝐩𝐭𝐢𝐨𝐧: ")
        rr=rr[0]
        rr=rr.strip()        
        print(rr)     
        conn = sqlite3.connect('cart.db')
        cursor=conn.execute("UPDATE COMPANY set name='{}',productID='{}',price='{}' where ID = {}".format(dd,ccv,rr,int(update.effective_user.id)))
        conn.commit()
        conn = sqlite3.connect('cart.db')
        cursor = conn.execute("SELECT price from COMPANY where id= '{}' ".format(update.effective_user.id))
        conn.commit()
        for row in cursor:
            ffa=float(row[0])
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT type from COMPANY where productID= '{}' ".format(ccv))
        conn.commit()
        for row in cursor:
            poda=row[0]

        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT balance,name from COMPANY where id= '{}' ".format(update.effective_user.id))
        conn.commit()
        for row in cursor:
            ff=float(row[0])
            tep=row[1]
            tola=float(ff)-float(ffa)
            if ff>=ffa:   
                if str(poda)=='physical':
                    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
                    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
                    context.bot.send_message(chat_id=update.effective_user.id,text='Kindly Provide your Delivery address',reply_markup=reply_markup)
                    return ADRS
                elif str(poda)=='digital':
                    cursor=conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(tola,int(update.effective_user.id)))
                    conn.commit() 
                    yu= random.randint (0,999999)
                    xg=er.today().strftime("%m/%d/%Y, %H:%M:%S")
                    conn = sqlite3.connect('orders.db')
                    conn.execute("INSERT INTO COMPANY (ID,price,oid,pname,username,date,productID,status,address) \
                                        VALUES ('{}', '{}','{}','{}','{}','{}','{}','{}','{}')".format(str(update.effective_user.id),ffa,yu,dd,tep,xg,ccv,'Pending','N/A'))
                    conn.commit()
                    context.bot.send_message(chat_id=6417933558,text='You have a new order ')
                    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
                    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
                    context.bot.send_message(chat_id=update.effective_user.id,text='Thank you for your order Admin will provide you delivery',reply_markup=reply_markup)
                    return BUTTON
            else:             
                keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
                context.bot.send_message(chat_id=update.effective_user.id,text="Your balance is not enough".format(update.effective_user.id),reply_markup=reply_markup)
                return BUTTON 
    elif a=='90':
        keyf=[]
        conn = sqlite3.connect('cat.db')
        cursor = conn.execute("SELECT cat from COMPANY")
        conn.commit()
        keyf=[]
        for row in cursor:
            c=[InlineKeyboardButton(row[0], callback_data=row[0])]
            keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True ) 

        context.bot.send_message(chat_id=update.effective_user.id,text="Select the category to delete",reply_markup=reply_markup)  
        return DELCAT
    elif a=="ff":    
        conn = sqlite3.connect('faq.db')
        cursor = conn.execute("SELECT question,code from COMPANY ")
        conn.commit()
        keyf=[]
        for row in cursor:
            c=[InlineKeyboardButton(row[0], callback_data=row[1])]
            keyf.append(c) 
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True ) 
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Question to read out",reply_markup=reply_markup)
        return FFAQ
    elif a== '100':
            keyboard =[[InlineKeyboardButton("➕ Product", callback_data="1"),InlineKeyboardButton("❌ Product", callback_data="2")],
                    [InlineKeyboardButton("Add category", callback_data="32"),InlineKeyboardButton("Delete Category", callback_data="90")],
                    [InlineKeyboardButton("🛒 Orders", callback_data="99v"),InlineKeyboardButton("🛒 Users Log", callback_data="llo")],
                    [InlineKeyboardButton("➕ FAQ", callback_data="1f"),InlineKeyboardButton("❌ FAQ", callback_data="2f")],
                    [InlineKeyboardButton("🔊 Announcement", callback_data="916"),InlineKeyboardButton("Shop logo", callback_data="8ab")],
                    [InlineKeyboardButton("User Side", callback_data="100"),InlineKeyboardButton("Add balance", callback_data="abda")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to admin Dashboard" ,reply_markup=reply_markup)
            return BUTTON
    elif a== '200':
        loc_keyboard1 = KeyboardButton(text="🛒 Store")
        loc_keyboard2 = KeyboardButton(text="💵 Wallet")
        loc_keyboard3 = KeyboardButton(text="🗓 Orders")
        loc_keyboard4 = KeyboardButton(text="☎️ Support")
        keyboard = [[loc_keyboard1,loc_keyboard3],[loc_keyboard2,loc_keyboard4]]
        reply_markup = ReplyKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_photo(chat_id=update.effective_user.id,photo=open('logo.jpg','rb'),caption=wc[0],reply_markup=reply_markup)
        return BUTTON 
    elif a== '916':
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send announcement message',reply_markup=reply_markup)
            return JE
    elif a=='8ab':   
            keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send shop logo',reply_markup=reply_markup)
            return LOG
    elif a=='abda':   
            keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send User ID to add balance',reply_markup=reply_markup)
            return KAKAK
    elif a== '1bal':
        keyboard =[[InlineKeyboardButton("Crypto", callback_data="crypto"),InlineKeyboardButton("Stripe", callback_data="stripe")],[InlineKeyboardButton("🔙 Back", callback_data="200")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Select Payment Method to add balance in your Wallet',reply_markup=reply_markup)
        return BUTTON 
    elif a== 'llo':
        style = xlwt.easyxf('font: bold 1, color black;')
        conn = sqlite3.connect('wallet.db') 
        cursor = conn.execute("SELECT ID,name from COMPANY ")
        conn.commit() 
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            f = open("log.txt", "w",encoding="utf-8")
            conn = sqlite3.connect('wallet.db') 
            cursor = conn.execute("SELECT ID,name,balance from COMPANY ")
            conn.commit()
            xa="log\n"
            i=0
            j=0
            workbook = xlwt.Workbook()
            sheet = workbook.add_sheet("Order History")
            sheet.write(i, 1, "user_id", style)
            sheet.write(i, 2, "user_name", style)
            sheet.write(i, 3, "Blance", style)
            for row in cursor:
                    i=i+1
                    j=j+1
                    inv=row[0]
                    vgy=row[1]
                    sheet.write(i, 1, row[0], style)
                    sheet.write(i, 2, row[1], style)
                    sheet.write(i, 3, row[2], style)
            workbook.save("log.xls") 
            keyboard =[[InlineKeyboardButton("Main Menu", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_document(chat_id=update.effective_user.id,document=open('log.xls', 'rb'),reply_markup=reply_markup)
        else:
            keyboard =[[InlineKeyboardButton("❌", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="You have no log",reply_markup=reply_markup)
            return BUTTON
    elif a=='120t':
        c=update.callback_query.message.message_id
        cc=update.callback_query.message.text
        d=cc#"\nSeller: "+row[10]+"\nStatus: "
        cc=cc.replace("$","")
        df=cc.split("OrderID: ")
        df=df[1]
        df=df.split("Product Name: ")
        df=df[0]
        df=df.strip()
        dd=cc.split("UserID:  ")
        dd=dd[1]
        dd=dd.split("OrderID: ")
        dd=dd[0]
        dd=dd.strip()

        ggg=cc.split("Price: ")
        ggg=ggg[1]
        ggg=ggg.split("User name:  ")
        ggg=ggg[0]
        ggg=ggg.strip()
        print(df) 
        global abbas
        abbas=df
        global abbase
        abbase=dd
        keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='Type and send delivery URL to user.\nOrderID: {}'.format(df),reply_markup=reply_markup)
        return AMAR
    elif a=='120':
        c=update.callback_query.message.message_id
        cc=update.callback_query.message.text
        d=cc#"\nSeller: "+row[10]+"\nStatus: "
        cc=cc.replace("$","")
        df=cc.split("OrderID: ")
        df=df[1]
        df=df.split("Product Name: ")
        df=df[0]
        df=df.strip()
        dd=cc.split("UserID:  ")
        dd=dd[1]
        dd=dd.split("OrderID: ")
        dd=dd[0]
        dd=dd.strip()

        ggg=cc.split("Price: ")
        ggg=ggg[1]
        ggg=ggg.split("User name:  ")
        ggg=ggg[0]
        ggg=ggg.strip()
        print(df)
        hhhh="Your order is being processed\n𝐎𝐫𝐝𝐞𝐫 𝐧𝐮𝐦𝐛𝐞𝐫# {}".format(df)  
        conn = sqlite3.connect("orders.db")  
        conn.execute("UPDATE COMPANY set  status='{}' where oid = '{}'".format('Accepted',df))
        conn.commit()
        conn.close() 
        context.bot.send_message(chat_id=dd,text=hhhh) 
        keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='accepted',reply_markup=reply_markup)
        return BUTTON 
    elif a=='130':
        c=update.callback_query.message.message_id
        cc=update.callback_query.message.text
        d=cc#"\nSeller: "+row[10]+"\nStatus: "
        cc=cc.replace("$","")
        df=cc.split("OrderID: ")
        df=df[1]
        df=df.split("Product Name: ")
        df=df[0]
        df=df.strip()
        dd=cc.split("UserID:  ")
        dd=dd[1]
        dd=dd.split("OrderID: ")
        dd=dd[0]
        dd=dd.strip()

        ggg=cc.split("Price: ")
        ggg=ggg[1]
        ggg=ggg.split("User name:  ")
        ggg=ggg[0]
        ggg=ggg.strip()
        print(df)
        hhhh="Your order is being rejected\n𝐎𝐫𝐝𝐞𝐫 𝐧𝐮𝐦𝐛𝐞𝐫# {}".format(df)  
        conn = sqlite3.connect("orders.db")  
        conn.execute("UPDATE COMPANY set  status='{}' where oid = '{}'".format('Rejected',df))
        conn.commit()
        conn.close() 
        context.bot.send_message(chat_id=dd,text=hhhh) 
        keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='Rejected',reply_markup=reply_markup)
        return BUTTON 
    elif a=='crypto':
      connection = sqlite3.connect("wallet.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT balance,name FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
      for names in cursor:
        inv=names[0]
        gsr=names[1]
        inv=float(inv)
        inv=round(inv,5)
        inv=str(inv)
        keyboard =[[InlineKeyboardButton("Top Up", callback_data="topup")],[InlineKeyboardButton("🔙 Back", callback_data="200")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Wallet ID:  "+str(update.effective_user.id)+"\n\nName:  "+gsr+"\nBalance:  "+"£"+inv+"\n\nPlease Send Your Btc to:\n\n`{}`\n\nAfter confirmation click on below button".format(wallo[0]),reply_markup=reply_markup,parse_mode=ParseMode.MARKDOWN)
    elif a=="topup":
        c=update.callback_query.message.message_id
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text="Please send your transcational hash to confirm the payment",reply_markup=reply_markup)
        return PAY
def ham(update,context):
    query = update.callback_query
    msg=query.data
    if msg=="Cancel":
            keyboard =[[InlineKeyboardButton("➕ Product", callback_data="1"),InlineKeyboardButton("❌ Product", callback_data="2")],
                    [InlineKeyboardButton("Add category", callback_data="32"),InlineKeyboardButton("Delete Category", callback_data="90")],
                    [InlineKeyboardButton("🛒 Orders", callback_data="99v"),InlineKeyboardButton("🛒 Users Log", callback_data="llo")],
                    [InlineKeyboardButton("➕ FAQ", callback_data="1f"),InlineKeyboardButton("❌ FAQ", callback_data="2f")],
                    [InlineKeyboardButton("🔊 Announcement", callback_data="916"),InlineKeyboardButton("Shop logo", callback_data="8ab")],
                    [InlineKeyboardButton("User Side", callback_data="100"),InlineKeyboardButton("Add balance", callback_data="abda")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to admin Dashboard" ,reply_markup=reply_markup)
            return BUTTON
    elif msg=="Pending":
            connection = sqlite3.connect('orders.db')
            cursor = connection.cursor()  
            cursor.execute("SELECT oid,username from COMPANY where status= '{}'  ".format("Pending"))
            jobs = cursor.fetchall()
            if len(jobs) !=0: 
                conn = sqlite3.connect('orders.db')
                cursor = conn.execute("SELECT oid,username from COMPANY where status= '{}'".format("Pending"))
                conn.commit()
                keyf=[]
                for row in cursor:
                    c=[InlineKeyboardButton("⏳ "+"{},{}".format(row[0],row[1]), callback_data=row[0])]
                    keyf.append(c)
                b=[InlineKeyboardButton("🔙Back", callback_data='🔙Back')] 
                keyf.append(b)
                reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True,one_time_keyboard=True)
                context.bot.send_message(chat_id=update.effective_user.id,text="Select Order to check detail",reply_markup=reply_markup) 
                return PENORD
            else:
                    keyboard =[[InlineKeyboardButton("Main Menu", callback_data="100")]]
                    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                    context.bot.send_message(chat_id=update.effective_user.id,text="You have no Pending orders",reply_markup=reply_markup)
                    return BUTTON         
   
    elif msg=="History":

            style = xlwt.easyxf('font: bold 1, color black;')
            conn = sqlite3.connect('orders.db') 
            cursor = conn.execute("SELECT ID,price,oid,pname,username,date,productID,status,address from COMPANY ")
            conn.commit() 
            jobs = cursor.fetchall()
            if len(jobs) !=0:
                f = open("orders.txt", "w",encoding="utf-8")
                conn = sqlite3.connect('orders.db') 
                cursor = conn.execute("SELECT ID,price,oid,pname,username,date,productID,status,address from COMPANY  ")
                conn.commit()
                xa="Orders\n"
                i=0
                j=0
                workbook = xlwt.Workbook()
                sheet = workbook.add_sheet("Order History")
                sheet.write(i, 1, "order_id", style)
                sheet.write(i, 2, "user_id", style)
                sheet.write(i, 3, "date",  style)
                sheet.write(i, 4, "Product Name",  style)
                sheet.write(i, 5, "Price",  style)
                sheet.write(i, 6, "Users_name",  style)
                sheet.write(i, 7, "address",  style)

                for row in cursor:
                        i=i+1
                        j=j+1
                        inv=row[0]
                        vgy=row[1]
                        xdrt=row[2]
                        sheet.write(i, 1, row[2], style)
                        sheet.write(i, 2, row[0], style)
                        sheet.write(i, 3, row[5],  style)
                        sheet.write(i, 4, row[3],  style)
                        sheet.write(i, 5, row[1],  style)
                        sheet.write(i, 6, row[4],  style)
                        sheet.write(i, 8, row[8],  style)
                workbook.save("orders.xls")
                keyboard =[[InlineKeyboardButton("Main Menu", callback_data="100")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                context.bot.send_document(chat_id=update.effective_user.id,document=open('orders.xls', 'rb'),reply_markup=reply_markup)
            else:
                    keyboard =[[InlineKeyboardButton("❌", callback_data="100")]]
                    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                    context.bot.send_message(chat_id=update.effective_user.id,text="You have no orders",reply_markup=reply_markup)
                    return BUTTON
def penord(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)
    if msg=="🔙Back":
        query.answer()  
        keyboard =[
                   [InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='Orders',reply_markup=reply_markup)
        return HAMA
    else:
        conn = sqlite3.connect('orders.db')
        cursor = conn.execute("SELECT productID from COMPANY where oid ='{}'".format(msg))
        conn.commit()
        for row in cursor: 
            gtrp=row[0]
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT type from COMPANY where productID ='{}'".format(str(gtrp)))
        conn.commit()
        for row in cursor: 
            xde=row[0]
            print(xde)
        conn = sqlite3.connect('orders.db')
        cursor = conn.execute("SELECT ID,price,oid,pname,username,date,productID,status,address,username from COMPANY where oid ='{}'".format(msg))
        conn.commit()
        for row in cursor: 
            if xde=='physical':
                g="\nUserID:  "+row[0]+"\nOrderID: "+row[2]+"\nProduct Name: "+row[3]+"\nPrice: "+row[1]+'$'+"\nUser name:  "+row[4]+"\nAddress:  "+row[8]+"\nStatus: "+row[7]+"\n\n"
                keyboard =[[InlineKeyboardButton("Delivered", callback_data="120"),InlineKeyboardButton("Reject", callback_data="130")],[InlineKeyboardButton("🔙 Cancel", callback_data="100")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
                context.bot.send_message(chat_id=update.effective_user.id,text="Pending order:\n"+g,reply_markup=reply_markup)
                return BUTTON
            elif xde=='digital':
                g="\nUserID:  "+row[0]+"\nOrderID: "+row[2]+"\nProduct Name: "+row[3]+"\nPrice: "+row[1]+'$'+"\nUser name:  "+row[4]+"\nAddress:  "+row[8]+"\nStatus: "+row[7]+"\n\n"
                keyboard =[[InlineKeyboardButton("Deliver", callback_data="120t"),InlineKeyboardButton("Reject", callback_data="130")],[InlineKeyboardButton("🔙 Cancel", callback_data="100")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
                context.bot.send_message(chat_id=update.effective_user.id,text="Pending order:\n"+g,reply_markup=reply_markup)
                return BUTTON        
def adrs(update,context):
    msg=update.message.text
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("SELECT price,productID,name from COMPANY where id= '{}' ".format(update.effective_user.id))
    conn.commit()
    for row in cursor:
        ffa=float(row[0])
        wdf=row[1]
        poop=row[2]
    conn = sqlite3.connect('wallet.db')
    cursor = conn.execute("SELECT balance,name from COMPANY where id= '{}' ".format(update.effective_user.id))
    conn.commit()
    for row in cursor:
        ff=float(row[0])
        tep=row[1]
        tola=float(ff)-float(ffa)
        if ff>=ffa:   
            cursor=conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(tola,int(update.effective_user.id)))
            conn.commit() 
            yu= random.randint (0,999999)
            xg=er.today().strftime("%m/%d/%Y, %H:%M:%S")
            conn = sqlite3.connect('orders.db')
            conn.execute("INSERT INTO COMPANY (ID,price,oid,pname,username,date,productID,status,address) \
                                VALUES ('{}', '{}','{}','{}','{}','{}','{}','{}','{}')".format(str(update.effective_user.id),ffa,yu,poop,tep,xg,wdf,'Pending',msg))
            conn.commit()
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Thank you for your order Admin will provide you delivery',reply_markup=reply_markup)
            context.bot.send_message(chat_id=6417933558,text='You have new order. Check Order section. Thanks')

            return BUTTON
def ffaq(update,context):#PDB,PDVM,PDV,SAL
    query = update.callback_query
    msg=query.data
    conn = sqlite3.connect('faq.db')
    cursor = conn.execute("SELECT question,answer from COMPANY where code= '{}' ".format(msg))
    conn.commit()
    for row in cursor:
        ff=row[0]
        rr=row[1]
        keyboard =[[InlineKeyboardButton("Back to FAQs", callback_data="ff")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Question: {}\n\nAnswer: {}".format(ff,rr),reply_markup=reply_markup)
        return BUTTON
def abd(update,context):
    msg=update.message.text
    global papa
    papa=msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Send answer of FAQs",reply_markup=reply_markup)
    return WALL
def dela(update,context):
    query = update.callback_query
    msg=query.data
    if msg=='🌎Main Menu':
        keyboard =[[InlineKeyboardButton("➕ FAQ", callback_data="1"),InlineKeyboardButton("FAQ", callback_data="4563")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to admin Dashboard" ,reply_markup=reply_markup)
        return BUTTON
    else:
        conn = sqlite3.connect('faq.db')
        cursor = conn.execute("DELETE  from COMPANY where code='{}'".format(msg))
        conn.commit()
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="FAQ deleted",reply_markup=reply_markup)
        return BUTTON
def wall(update,context):
    msg=update.message.text
    yu= random.randint (0,999999)
    conn = sqlite3.connect('faq.db')
    conn.execute("INSERT INTO COMPANY (question,answer,code) \
                        VALUES ('{}','{}','{}')".format(papa,msg,yu))
    conn.commit()  
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="FAQ added",reply_markup=reply_markup)
    return BUTTON
def amar(update,context):
    msg=update.message.text
    hhhh="Your order is being processed\n𝐎𝐫𝐝𝐞𝐫 𝐧𝐮𝐦𝐛𝐞𝐫# {}".format(abbas)  
    conn = sqlite3.connect("orders.db")  
    conn.execute("UPDATE COMPANY set  status='{}' where oid = '{}'".format('Accepted',abbas))
    conn.commit()
    conn.close() 
    context.bot.send_message(chat_id=6417933558,text='You have a new order')
    context.bot.send_message(chat_id=abbase,text=hhhh)
    context.bot.send_message(chat_id=abbase,text="Here is your delivery link. Open to download"+msg)
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Message delivered to the user",reply_markup=reply_markup)
    return BUTTON
def jos(update,context):
    msg=update.message.text
    print(msg)
    if msg=='🛒 Store':
        conn = sqlite3.connect('cat.db')
        cursor = conn.execute("SELECT cat from COMPANY")
        conn.commit()
        keyf=[] 
        for row in cursor:
            c=[InlineKeyboardButton(row[0], callback_data=row[0])]
            keyf.append(c) 
        b=[InlineKeyboardButton("Back", callback_data='🌎Main Menu')]
        keyf.append(b) 
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True ) 
        context.bot.send_message(chat_id=update.effective_user.id,text="Select category to view the products.",reply_markup=reply_markup)
        return PDM
    elif msg=='💵 Wallet':#t0J60jhxFVBjTywgbu7kCMLJZNdqhfSzVFBQAhjRzrQ
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT balance,tref from COMPANY where ID= {} ".format(int(update.effective_user.id)))
        conn.commit()
        for row in cursor:
            hj=float(row[0])
            hj=round(hj, 2)
            keyboard =[[InlineKeyboardButton("Add Balance", callback_data="1bal"),InlineKeyboardButton("🔙 Back", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='WalletID: {}\nBalance: {}$'.format(str(update.effective_user.id),row[0]),reply_markup=reply_markup)
            return BUTTON 
    elif msg=="🗓 Orders":
        conn = sqlite3.connect('orders.db')
        cursor = conn.execute("SELECT ID,price,oid,pname,username,date,productID,date,address,status from COMPANY where ID ='{}'".format(update.effective_user.id))
        conn.commit()
        for row in cursor: 
            g="\nUserID:  "+row[0]+"\nOrderID: "+row[2]+"\nProduct Name: "+row[3]+"\nPrice: "+row[1]+'$'+"\nUser name:  "+row[4]+"\nAddress:  "+row[8]+"\nDate: "+row[7]+"\nStatus: "+row[9]+"\n\n"
            keyboard =[[InlineKeyboardButton("🔙 Cancel", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text=g,reply_markup=reply_markup)
        return BUTTON  
    elif msg=='☎️ Support':
            keyboard =[[InlineKeyboardButton("📜 FAQs", callback_data="ff"),InlineKeyboardButton("Contact us", url='https://t.me/bonico1')]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Please Read through FAQ’S before sending message to admin.\n\n",reply_markup=reply_markup)
            return BUTTON
def pdm(update,context):#PDB,PDVM,PDV,SAL
    query = update.callback_query
    msg=query.data
    print(msg)    
    c=update.callback_query.message.message_id
    context.bot.delete_message(chat_id=update.effective_user.id,
                        message_id=c)
    if msg=="🌎Main Menu":
        loc_keyboard1 = KeyboardButton(text="🛒 Store")
        loc_keyboard2 = KeyboardButton(text="💵 Wallet")
        loc_keyboard3 = KeyboardButton(text="🗓 Orders")
        loc_keyboard4 = KeyboardButton(text="☎️ Support")
        keyboard = [[loc_keyboard1,loc_keyboard3],[loc_keyboard2,loc_keyboard4]]
        reply_markup = ReplyKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_photo(chat_id=update.effective_user.id,photo=open('logo.jpg','rb'),caption=wc[0],reply_markup=reply_markup)
        return BUTTON 
    elif msg=="200":
        loc_keyboard1 = KeyboardButton(text="🛒 Store")
        loc_keyboard2 = KeyboardButton(text="💵 Wallet")
        loc_keyboard3 = KeyboardButton(text="🗓 Orders")
        loc_keyboard4 = KeyboardButton(text="☎️ Support")
        keyboard = [[loc_keyboard1,loc_keyboard3],[loc_keyboard2,loc_keyboard4]]
        reply_markup = ReplyKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_photo(chat_id=update.effective_user.id,photo=open('logo.jpg','rb'),caption=wc[0],reply_markup=reply_markup)
        return BUTTON 
    else:
        conn = sqlite3.connect('wallet.db')
        cursor=conn.execute("UPDATE COMPANY set ct = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
        conn.commit()
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT name from COMPANY where category= '{}' ".format(msg))  
        conn.commit()                             
        keyf=[]
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('shop.db')
            cursor = conn.execute("SELECT name from COMPANY where category= '{}' ".format(msg))
            conn.commit()
            for row in cursor:
                c=[InlineKeyboardButton(row[0], callback_data=row[0])]
                keyf.append(c)
            b=[InlineKeyboardButton("Back", callback_data='🌎Main Menu')]
            keyf.append(b)
            reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text="🛒 Shop\n\nSelect Name of Product to View ",reply_markup=reply_markup)            
            return PDB
        else: 
            keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text='This Category has no Items',reply_markup=reply_markup)
            return BUTTON
def pdb(update,context):#PDB,PDVM,PDV,SAL
    query = update.callback_query
    msg=query.data
    print(msg)    
    c=update.callback_query.message.message_id
    context.bot.delete_message(chat_id=update.effective_user.id,
                        message_id=c)
    if msg=="🌎Main Menu":
        conn = sqlite3.connect('cat.db')
        cursor = conn.execute("SELECT cat from COMPANY")
        conn.commit()
        keyf=[] 
        for row in cursor:
            c=[InlineKeyboardButton(row[0], callback_data=row[0])]
            keyf.append(c) 
        b=[InlineKeyboardButton("Back", callback_data='🌎Main Menu')]
        keyf.append(b) 
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True ) 
        context.bot.send_message(chat_id=update.effective_user.id,text="Select category to view the products.",reply_markup=reply_markup)
        return PDM
    elif msg=="200":
        loc_keyboard1 = KeyboardButton(text="🛒 Store")
        loc_keyboard2 = KeyboardButton(text="💵 Wallet")
        loc_keyboard3 = KeyboardButton(text="🗓 Orders")
        loc_keyboard4 = KeyboardButton(text="☎️ Support")
        keyboard = [[loc_keyboard1,loc_keyboard3],[loc_keyboard2,loc_keyboard4]]
        reply_markup = ReplyKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_photo(chat_id=update.effective_user.id,photo=open('logo.jpg','rb'),caption=wc[0],reply_markup=reply_markup)
        return BUTTON 
    else:
            conn = sqlite3.connect('wallet.db')
            cursor = conn.execute("SELECT ct from COMPANY where ID={} ".format(str(update.effective_user.id)))
            conn.commit()
            for names in cursor:
                hjj=names[0] 
            conn = sqlite3.connect('shop.db')
            cursor = conn.execute("SELECT name from COMPANY where name='{}' AND category= '{}'".format(msg,hjj))                              
            keyf=[]
            jobs = cursor.fetchall()
            if len(jobs) !=0:
                conn = sqlite3.connect('shop.db')
                cursor = conn.execute("SELECT name,price,productID,image,description from COMPANY where name='{}' AND category= '{}'".format(msg,hjj))  
                for row in cursor:                               
                    m="𝐒𝐞𝐫𝐯𝐢𝐜𝐞 𝐍𝐚𝐦𝐞: {}\n𝐏𝐫𝐨𝐝𝐮𝐜𝐭𝐈𝐃: {}\n𝐏𝐫𝐢𝐜𝐞: {}$\n𝐃𝐞𝐬𝐜𝐫𝐢𝐩𝐭𝐢𝐨𝐧: {}".format(row[0],row[2],row[1],row[4])

                    keyboard =[[InlineKeyboardButton("Buy Now", callback_data="buy")],[InlineKeyboardButton("🔙 Back", callback_data="200d")]]
                    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                    context.bot.send_photo(chat_id=update.effective_user.id,photo=row[3],caption=m,reply_markup=reply_markup)
                    return BUTTON
            else:
                keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="200")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
                context.bot.send_message(chat_id=update.effective_user.id,text='This Category has no Items',reply_markup=reply_markup)
                return BUTTON 
def log(update,context):
    msg=update.message.photo[-1].file_id
    newFile = context.bot.getFile(msg)
    newFile.download('logo.jpg')
    keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text='Logo edited successfully',reply_markup=reply_markup)
    return BUTTON
def je(update,context):
        msg=update.message.text
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT ID from COMPANY")
        conn.commit()
        for row in cursor:
            aa=row[0]
            try:
                context.bot.send_message(chat_id=aa,text=msg)
            except:
                pass
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Message Sent",reply_markup=reply_markup)
        return BUTTON
def delcat(update,context):
    c=update.callback_query.message.message_id
    query = update.callback_query
    msg=query.data
    conn = sqlite3.connect('cat.db')
    cursor = conn.execute("DELETE from COMPANY where cat='{}'".format(msg))
    conn.commit()
    keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Category Deleted",reply_markup=reply_markup)
    return BUTTON
def mcat(update,context):
    msg=update.message.text
    conn = sqlite3.connect('cat.db')
    conn.execute("INSERT INTO COMPANY (cat) \
                        VALUES ('{}')".format(msg))
    conn.commit()

    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="category added successfully" ,reply_markup=reply_markup)
    return BUTTON
def delete(update,context):
    msg=update.message.text
    conn = sqlite3.connect('shop.db')
    cursor = conn.execute("SELECT productID from COMPANY  where productID = '{}'".format(str(msg)))
    conn.commit()
    jobs = cursor.fetchall()
    if len(jobs) !=0:
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("DELETE  from COMPANY where productID='{}'".format(msg))
        conn.commit()

        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Product Deleted",reply_markup=reply_markup)
        return BUTTON
    else:
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Invalid ProductID",reply_markup=reply_markup)
        return DELETE
def kakak(update,context):
    msg=update.message.text
    global xtar
    xtar=msg
    conn = sqlite3.connect('wallet.db')
    cursor = conn.execute("SELECT ID from COMPANY  where ID = '{}'".format(msg))
    conn.commit()
    jobs = cursor.fetchall()
    if len(jobs) !=0:

        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Send Balance in digit only",reply_markup=reply_markup)
        return KUNG
    else:
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Invalid ID",reply_markup=reply_markup)
        return KAKAK
def kung(update,context):
    msg=update.message.text
    try:
        msg=float(msg)
        connection = sqlite3.connect("wallet.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(xtar)) )
        for names in cursor:
            inv=float(names[0])
        bn=inv+msg
        bn=str(bn)
        conn = sqlite3.connect("wallet.db")  
        conn.execute("UPDATE COMPANY set bala'99vnce = '{}' where ID = {}".format(bn,int(xtar)))
        conn.commit()
        conn.close()
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Balance added",reply_markup=reply_markup)
        return BUTTON
    except:
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Send balance in digits only",reply_markup=reply_markup)
        return KUNG
def ab(update,context):
    msg=update.message.text 
    global nm
    nm =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Send Product Price.',reply_markup=reply_markup)
    return PPR
def ppr(update,context):
    msg=update.message.text 
    global pri
    pri =msg
    try:
        msg=float(msg)
        keyf=[]
        conn = sqlite3.connect('cat.db')
        cursor = conn.execute("SELECT cat from COMPANY")
        conn.commit()
        keyf=[]
        for row in cursor:
            c=[InlineKeyboardButton(row[0], callback_data=row[0])]
            keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True ) 
        context.bot.send_message(chat_id=update.effective_user.id,text="Select the Category for product",reply_markup=reply_markup)  
        return CATE
    except:
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send Product Price in digits only.',reply_markup=reply_markup)
        return PPR
def cate(update,context):
    query = update.callback_query
    msg=query.data
    global categ
    categ =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Send Description of Product.',reply_markup=reply_markup)
    return DESCRIP
def descrip(update,context):
    msg=update.message.text 
    global desa
    desa =msg
    keyboard =[[InlineKeyboardButton("Physical", callback_data="physical"),InlineKeyboardButton("Digital", callback_data="digital")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Select product type.',reply_markup=reply_markup)
    return PTYPE
def ptype(update,context):
    query = update.callback_query
    msg=query.data
    global ptop
    ptop =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Send Image of Product.',reply_markup=reply_markup)
    return IMGE
def imge(update,context):
    msg=update.message.photo[-1].file_id
    yu= random.randint (0,999999)
    conn = sqlite3.connect('shop.db')
    conn.execute("INSERT INTO COMPANY (name,price,image,productID,category,description,type) \
                        VALUES ('{}', '{}','{}','{}','{}','{}','{}')".format(nm,pri,msg,yu,categ,desa,ptop))
    conn.commit()

    keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text='Product added successfully',reply_markup=reply_markup)
    return BUTTON
def pay(update,context):
    msg=update.message.text
    connection = sqlite3.connect("hash.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT hash FROM COMPANY where hash= '{}'".format(msg)) 
    jobs = cursor.fetchall()
    if len(jobs) ==0:
        transaction_hash = msg
        transaction_details = get_transaction_details(transaction_hash)
        try:
            if transaction_details:
                btc_amount = sum([input['prev_out']['value'] for input in transaction_details['inputs']]) / 1e8
                usd_equivalent = btc_amount * get_current_btc_to_usd_rate()
                usd_equivalent=round(usd_equivalent,2)
                connection = sqlite3.connect("wallet.db")  
                cursor = connection.cursor()  
                cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
                for names in cursor:
                  inv=float(names[0])
                bn=inv+usd_equivalent
                bn=str(bn)
                conn = sqlite3.connect("wallet.db")  
                conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bn,int(update.effective_user.id)))
                conn.commit()
                conn.close()
                conn = sqlite3.connect('hash.db')
                conn.execute("INSERT INTO COMPANY (ID,hash) \
                        VALUES ('{}','{}')".format(update.effective_user.id,msg)) 
                conn.commit()
                keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                context.bot.send_message(chat_id=update.effective_user.id,text="{}£ is added to your account".format(usd_equivalent),reply_markup=reply_markup)


            else:
                keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                context.bot.send_message(chat_id=update.effective_user.id,text="Transcation not found yet wait for confirmation",reply_markup=reply_markup)
        except:
          keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
          context.bot.send_message(chat_id=update.effective_user.id,text="Transcation not found yet wait for confirmation",reply_markup=reply_markup)
       
       
    else: 
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="transcation already top up",reply_markup=reply_markup)

def admin(update,context):
    print(update.effective_user.id)
    user = update.message.from_user
    usa=str(update.effective_user.id)
   
    if usa == "80751603v0" or usa == "6417933558" or usa== "1394902938" or usa in man: 
            keyboard =[[InlineKeyboardButton("➕ Product", callback_data="1"),InlineKeyboardButton("❌ Product", callback_data="2")],
                    [InlineKeyboardButton("Add category", callback_data="32"),InlineKeyboardButton("Delete Category", callback_data="90")],
                    [InlineKeyboardButton("🛒 Orders", callback_data="99v"),InlineKeyboardButton("🛒 Users Log", callback_data="llo")],
                    [InlineKeyboardButton("➕ FAQ", callback_data="1f"),InlineKeyboardButton("❌ FAQ", callback_data="2f")],
                    [InlineKeyboardButton("🔊 Announcement", callback_data="916"),InlineKeyboardButton("Shop logo", callback_data="8ab")],
                    [InlineKeyboardButton("User Side", callback_data="100"),InlineKeyboardButton("Add balance", callback_data="abda")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to admin Dashboard" ,reply_markup=reply_markup)
            return BUTTON
   
def main():#AB,PPR,CATE,DESCRIP,IMGE
  updater = Updater("6320138661:AAGyEve-EWrkl-JlAWQucUXc7zEV32s4iVo", use_context=True)
  dp = updater.dispatcher
  dp.add_handler(CommandHandler("admin", admin))
  conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start),MessageHandler(Filters.regex('^(🔗 Wholesale|👤 Seller|🛒 Store|💵 Wallet|🗓 Orders|☎️ Support|🔈 Channel|💬 Channel)$'), jos),CallbackQueryHandler(button)],

        states={        
        BUTTON: [CallbackQueryHandler(button)],
        CATE: [CallbackQueryHandler(cate)],
        PDM: [CallbackQueryHandler(pdm)],
        PDB: [CallbackQueryHandler(pdb)],
        PTYPE: [CallbackQueryHandler(ptype)],
        AB: [MessageHandler(Filters.text, ab),CallbackQueryHandler(button)],
        PPR: [MessageHandler(Filters.text, ppr),CallbackQueryHandler(button)],
        DESCRIP: [MessageHandler(Filters.text, descrip),CallbackQueryHandler(button)],
        IMGE: [MessageHandler(Filters.photo, imge),CallbackQueryHandler(button)],
        DELETE: [MessageHandler(Filters.text, delete),CallbackQueryHandler(button)],
        MCAT: [MessageHandler(Filters.text, mcat),CallbackQueryHandler(button)],
        WALL: [MessageHandler(Filters.text, wall),CallbackQueryHandler(button)],
        ABD: [MessageHandler(Filters.text, abd),CallbackQueryHandler(button)],
        DELCAT: [CallbackQueryHandler(delcat)],
        DELA: [CallbackQueryHandler(dela)],
        PAY: [MessageHandler(Filters.text, pay),CallbackQueryHandler(button)],
        PENORD: [CallbackQueryHandler(penord)],
        JE: [MessageHandler(Filters.text, je),CallbackQueryHandler(button)],
        AMAR: [MessageHandler(Filters.text, amar),CallbackQueryHandler(button)],
        HAM: [CallbackQueryHandler(ham)],
        ADRS: [MessageHandler(Filters.text, adrs),CallbackQueryHandler(button)],
        FFAQ: [MessageHandler(Filters.text, ffaq),CallbackQueryHandler(button)],
        KAKAK: [MessageHandler(Filters.text, kakak),CallbackQueryHandler(button)],
        KUNG: [MessageHandler(Filters.text, kung),CallbackQueryHandler(button)],
        LOG: [MessageHandler(Filters.photo, log),CallbackQueryHandler(button)],

        },  
        fallbacks=[CommandHandler('start', start)],#KUNG,KAKAK
        allow_reentry=True
     )
  dp.add_handler(conv_handler)
  updater.start_polling()
  updater.idle()
    
if __name__ == '__main__':
    main()
